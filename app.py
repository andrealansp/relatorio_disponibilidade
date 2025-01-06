import locale
import logging
import os
import sys
import time
from datetime import datetime
from typing import List

import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait

from config import *
from emailsender import Emailer

logger = logging.getLogger(__name__)
logging.basicConfig(filename='app.log', format='%(asctime)s - %(levelname)s: %(message)s', encoding='utf-8',
                    level=logging.DEBUG)


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# Inicializa variáveis para acesso aos diretórios.
diretorio_downloads: str = resource_path("downloads")
diretorio_disponibilidade: str = resource_path("disponibilidade")

# Inicializa instancia do navegador.
navegador = None


# Remove o arquivo status.xlsx baixado anteriormente
def remove_arquivos():
    try:
        global diretorio_downloads
        # configurando diretorios
        arquivos: List[str] = os.listdir(diretorio_downloads)

        # Remove arquivo anterior
        for arquivo in arquivos:
            os.remove(os.path.join(diretorio_downloads, arquivo))
    except FileNotFoundError:
        logging.error(f"{FileNotFoundError.__str__()} - Sem arquivos na pasta")


# Inicia o driver do navegador
def iniciar_driver():
    chrome_options = Options()

    arguments = ["--lang=en-US", "--window-size=1300,1000", "--headless"]

    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": diretorio_downloads,
            "safebrowsing.enabled": "false",
            "download.prompt_for_download": False,
            "profile.default_content_setting_values.notifications": 2,
            "profile.default_content_setting_values.automatic_downloads": 1,
        },
    )

    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(
        driver,
        30,
        poll_frequency=1,
        ignored_exceptions=[
            NoSuchElementException,
            ElementNotVisibleException,
            ElementNotSelectableException,
        ],
    )
    return driver, wait


# Realiza o login no site do dss
def logar():
    try:
        global navegador
        navegador, wait = iniciar_driver()
        navegador.get(SITE_DSS)
        time.sleep(20)
        campo_usuario = navegador.find_element(
            "xpath", "/html/body/div/div/div[3]/div[2]/div[1]/div/input"
        )
        campo_usuario.click()
        campo_usuario.send_keys(LOGIN_DAHUA)
        campo_senha = navegador.find_element(
            "xpath", "/html/body/div/div/div[3]/div[2]/div[2]/div[1]/input"
        )
        campo_senha.click()
        time.sleep(5)
        senha = PASSWORD_DAHUA

        for letra in senha:
            campo_senha.send_keys(letra)
            time.sleep((1 / 5))

        time.sleep(5)
        entrar = navegador.find_element("xpath", "/html/body/div/div/div[3]/div[2]/button")
        entrar.click()
        time.sleep(5)
    except NoSuchElementException:
        logging.error(f"{NoSuchElementException.__str__()}")


# Abre a área de dispositivos do site DSS
def acessar_dispositivos():
    try:
        link_dispositivo = navegador.find_element(
            "xpath", "/html/body/div/div/div[3]/div/div/div/div/div/div[1]/div/h3"
        )
        link_dispositivo.click()
        time.sleep(5)
    except NoSuchElementException:
        logging.error({NoSuchElementException.__str__()})


# Realiza o download do arquivo.
def realiza_download():
    try:
        link_exportado = navegador.find_element(
            "xpath",
            "/html/body/div/div/div[3]/div/div/div[2]/div[2]/div[1]/div[4]/span[2]/div/div[1]/a",
        )
        ActionChains(navegador).move_to_element(link_exportado).perform()
        link_conten_sub = navegador.find_element(
            "xpath",
            '//*[@id="page-content"]/div[2]/div[2]/div[1]/div[4]/span[2]/div/div[2]/div/div/ul/li[1]',
        )

        acoes = ActionChains(navegador)
        acoes.move_to_element(link_conten_sub)
        acoes.click(link_conten_sub)
        acoes.perform()
        time.sleep(20)
        navegador.quit()
    except NoSuchElementException:
        logging.error({NoSuchElementException.__str__()})


# Altera o nome da WorkSheet para status, a planilha original vem um nome composto pela data.
def renomeia_sheet():
    arquivos_baixados: list[str] = os.listdir(diretorio_downloads)
    try:
        for arquivo_baixado in arquivos_baixados:
            os.rename(
                os.path.join(diretorio_downloads, arquivo_baixado),
                os.path.join(diretorio_downloads, "status.xlsx"),
            )

        ss = openpyxl.load_workbook(os.path.join(diretorio_downloads, "status.xlsx"))
        tabela = ss.active
        tabela.title = "status"
        ss.save(os.path.join(diretorio_downloads, "status.xlsx"))
    except Exception as e:
        logging.error({e.__str__()})


# Calcula quantidade de câmeras online e quantidade de câmeras
def calcula_status():
    caminho_status = os.path.join(diretorio_downloads, "status.xlsx")
    caminho_regioes = os.path.join(diretorio_downloads, "regioes.xlsx")
    try:
        df = pd.read_excel(caminho_status)
        df_regioes = df.groupby("Organization Name")["Channel Status"].agg(['sum', 'count', "mean"])
        df_regioes_perkons = df_regioes.drop("FACIAL")
        regioes_renomeado = df_regioes_perkons.rename(
            columns={"sum": "Online", "count": "Offline", "mean": "Média"})
        regioes_renomeado["Média"] = regioes_renomeado["Média"] * 100
        regioes_renomeado.to_excel(caminho_regioes)

        cameras_online = df[(df['Organization Name'] != "FACIAL") &
                            (df['Organization Name'] != "OBRAS E VANDALISMOS")]\
                            ["Channel Status"].sum()

        media_disponibilidade = \
            df[(df['Organization Name'] != "FACIAL") &
               (df['Organization Name'] != "OBRAS E VANDALISMOS")]\
                ["Channel Status"].mean() * 100

        quantidade_cameras = \
            df[(df['Organization Name'] != "FACIAL") &
               (df['Organization Name'] != "OBRAS E VANDALISMOS")]\
                ["Channel Status"].count()

    except Exception as e:
        logging.error({e.__str__()})
    return quantidade_cameras, cameras_online, media_disponibilidade


# Atualiza o arquivo de disponibilidade
def atualizar_disponibilidade():
    try:
        caminho_disponiblidade = os.path.join(
            diretorio_disponibilidade, "DISPONIBILIDADE_CONTRATO.xlsx"
        )
        disponibilidade = openpyxl.load_workbook(caminho_disponiblidade, data_only=True)
        # definindo variáveis importante
        locale.setlocale(locale.LC_ALL, "pt_br")
        data_de_hoje = datetime.now()
        str_data_de_hoje = data_de_hoje.strftime("%Y-%m-%d 00:00:00")
        mes_numero = data_de_hoje.month
        sheet_mes_atual = disponibilidade.active
        quantidade_cameras, cameras_online, media_disponibilidade = calcula_status()
        for linha in sheet_mes_atual.iter_rows(min_row=2):
            data = str(linha[0].value)
            if str_data_de_hoje == data:
                # Preenchendo as três colunas com os dados calculado na planilha status
                # AMOSTRA1 CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE
                if linha[1].value is None:
                    linha[1].value = cameras_online
                    linha[2].value = quantidade_cameras
                    linha[3].value = media_disponibilidade
                    linha[19].value = linha[3].value

                    logging.INFO("Amostra 1 - Realizada")
                    enviar_email_robo(data_de_hoje, cameras_online, quantidade_cameras, linha[19].value)
                    break

                # AMOSTRA2  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE
                if linha[4].value is None:
                    linha[4].value = cameras_online
                    linha[5].value = quantidade_cameras
                    linha[6].value = media_disponibilidade
                    linha[19].value = round(
                        ((float(linha[3].value) + float(linha[6].value)) / 2), 4
                    )
                    logging.INFO("Amostra 2 - Realizada")
                    enviar_email_robo(data_de_hoje, cameras_online, quantidade_cameras, linha[19].value)
                    break

                # AMOSTRA3  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE
                if linha[7].value is None:
                    linha[7].value = cameras_online
                    linha[8].value = quantidade_cameras
                    linha[9].value = media_disponibilidade
                    linha[19].value = round(
                        (
                                (
                                        float(linha[3].value)
                                        + float(linha[6].value)
                                        + float(linha[9].value)
                                )
                                / 3
                        ),
                        4,
                    )
                    logging.INFO("Amostra 3 - Realizada")
                    enviar_email_robo(data_de_hoje, cameras_online, quantidade_cameras, linha[19].value)
                    break

                # AMOSTRA4  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE
                if linha[10].value is None:
                    linha[10].value = cameras_online
                    linha[11].value = quantidade_cameras
                    linha[12].value = media_disponibilidade
                    linha[19].value = round(
                        (
                                (
                                        float(linha[3].value)
                                        + float(linha[6].value)
                                        + float(linha[9].value)
                                        + float(linha[12].value)
                                )
                                / 4
                        ),
                        4,
                    )
                    logging.INFO("Amostra 4 - Realizada")
                    enviar_email_robo(data_de_hoje, cameras_online, quantidade_cameras, linha[19].value)
                    break

                # AMOSTRA5  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE
                if linha[13].value is None:
                    linha[13].value = cameras_online
                    linha[14].value = quantidade_cameras
                    linha[15].value = media_disponibilidade
                    linha[19].value = round(
                        (
                                (
                                        float(linha[3].value)
                                        + float(linha[6].value)
                                        + float(linha[9].value)
                                        + float(linha[12].value)
                                        + float(linha[15].value)
                                )
                                / 5
                        ),
                        4,
                    )
                    logging.INFO("Amostra 5 - Realizada")
                    enviar_email_robo(data_de_hoje, cameras_online, quantidade_cameras, linha[19].value)
                    break

                # AMOSTRA6  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE
                if linha[16].value is None:
                    linha[16].value = cameras_online
                    linha[17].value = quantidade_cameras
                    linha[18].value = media_disponibilidade
                    linha[19].value = round(
                        (
                                (
                                        float(linha[3].value)
                                        + float(linha[6].value)
                                        + float(linha[9].value)
                                        + float(linha[12].value)
                                        + float(linha[15].value)
                                        + float(linha[18].value)
                                )
                                / 6
                        ),
                        4,
                    )
                    logging.INFO("Amostra 6 - Realizada")
                    enviar_email_robo(data_de_hoje, cameras_online, quantidade_cameras, linha[19].value)
                    break
    except Exception as e:
        logging.error({e.__str__()})

    disponibilidade.save(caminho_disponiblidade)


def checar_none(valor):
    if valor is None:
        return 0
    else:
        return valor


def retorna_dados_regiao():
    try:
        caminho_disponiblidade_regioes = os.path.join(
            diretorio_downloads, "regioes.xlsx"
        )
        wb = openpyxl.load_workbook(caminho_disponiblidade_regioes)
        ws = wb.active
        dados_regioes = ""
        for linha in ws.iter_rows(min_row=2):
            dados_regioes += f"""
            <tr>
            <td style="width:150px text-align="center";">{linha[0].value}</td>
            <td style="width:150px text-align="center";">{linha[1].value}</td>
            <td style="width:150px text-align="center";">{linha[2].value}</td>
            <td style="width:150px text-align="center";">{round(float(linha[3].value), 2)}%</td>
            </tr>
            """
    except Exception as e:
        logging.error({e.__str__()})

    return dados_regioes


def enviar_email_robo(data, online, total, media) -> None:
    dados_regioes = retorna_dados_regiao()
    email = Emailer(EMAIL_ADDRESS, EMAIL_PASSWORD)
    email.definir_conteudo(
        topico="Robô disponibilidade",
        email_remetente=EMAIL_ADDRESS,
        lista_contatos=[
            "a.alves@perkons.com",
            "fernando.b@perkons.com",
            "alexander.s@perkons.com",
        ],
        conteudo_email=f"""<h3>Disponibilidade do Sistema:</h3> <p>Data: {data.strftime("%d/%m/%y - %H:%M")}</br> 
                                Quantidade de Câmeras Online: {online}</br>
                                Quantidade de Câmeras no Total: {total}</br>
                                Disponibilidade em Porcentagem: {media}% </p>
                                
                                <h3>Disponibilidade por regiao:</h3>
                                <table>
                                    <thead>
                                    <tr>
                                        <th style="width:150px; text-align="center";>Região</th>
                                        <th style="width:150px; text-align="center";>Online</th>
                                        <th style="width:150px; text-align="center";>Total</th>
                                        <th style="width:150px; text-align="center";>Média</th>
                                    </tr>
                                    </thead>
                                    <tbody>
                                    {dados_regioes}
                                    </tbody>    
                                </table>
                                
                                """
    )
    email.enviar_email(5)


remove_arquivos()
logar()
acessar_dispositivos()
realiza_download()
renomeia_sheet()
atualizar_disponibilidade()
